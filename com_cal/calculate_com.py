# -*- coding: utf-8 -*-
"""根据姿态关键点计算人体质心(CoM)并在视频帧上绘制叠加标记的脚本。

整体流程：
    1. 从 Excel(``视频标注box.xlsx``)中读取某位患者每个视频片段的元数据
       (文件夹名称、视频文件名、起始/结束帧号)。
    2. 在患者目录下扫描所有 ``all_frames_ankle_data.json``，按
       (会话名, 归一化视频名, 起始帧, 结束帧) 建立索引。
    3. 把 Excel 中的元数据与磁盘上的 JSON 片段一一匹配。
    4. 对每个匹配上的片段，逐帧用 7 段人体测量学模型计算质心坐标。
    5. 把每帧质心结果写出为 ``*_COM.json``，并在原始帧图上绘制质心圆点叠加图。

坐标说明：这里的关键点是 2D 像素坐标(原点在左上角，Y 轴向下)。
"""

from __future__ import annotations

import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# cv2(OpenCV)仅在绘制叠加图时才需要；这里做软导入，
# 缺失时不直接报错，等真正调用绘图函数时再抛出明确提示。
try:
    import cv2
except ImportError:  # pragma: no cover - runtime dependency check
    cv2 = None


# 所有相机数据的根目录，以及记录视频标注信息的 Excel 路径。
CAMERA_DATA_ROOT = Path(r"H:\Camera_data")
EXCEL_PATH = CAMERA_DATA_ROOT / "视频标注box.xlsx"


# 7 段人体测量学模型定义，每个元素为：
#   (段名称, 近端关键点, 远端关键点, 该段质量占全身比例, 近端->远端的质心位置比例)
# 关键点用 COCO-17 索引表示；若为元组(如 (11,12))表示取两个关键点的中点作为锚点。
# 质量比例与质心比例参考人体测量学常数(trunk 段最重，占 0.578)。
SEGMENTS = [
    ("trunk_head_neck", (11, 12), (5, 6), 0.578, 0.660),
    ("left_total_arm", 5, 9, 0.050, 0.530),
    ("right_total_arm", 6, 10, 0.050, 0.530),
    ("left_foot_and_leg", 13, 15, 0.061, 0.606),
    ("right_foot_and_leg", 14, 16, 0.061, 0.606),
    ("left_thigh", 11, 13, 0.100, 0.433),
    ("right_thigh", 12, 14, 0.100, 0.433),
]


@dataclass(frozen=True)
class SegmentMeta:
    """单个视频片段的元数据(来自 Excel 一行)。"""

    session_name: str   # 会话/文件夹名称
    video_name: str     # 视频文件名(含扩展名)
    video_stem: str     # 视频文件名(不含扩展名)
    start_frame: int    # 该片段起始帧号
    end_frame: int      # 该片段结束帧号


@dataclass(frozen=True)
class SegmentSource:
    """磁盘上某个片段对应的数据来源(JSON 文件与所在目录)。"""

    json_path: Path     # all_frames_ankle_data.json 的完整路径
    segment_dir: Path   # 该 JSON 所在目录(同时也是原始帧图所在目录)


def normalize_video_name(value: str) -> str:
    """把视频名归一化：去掉所有非字母数字字符并转小写。

    用于匹配 Excel 中的视频名与磁盘目录名(两者命名风格可能不同)。
    """
    return re.sub(r"[^A-Za-z0-9]", "", value).lower()


def parse_int(value: Any) -> int:
    """把各种类型的"类整数"值稳健地解析为 int。

    处理 None、int、float，以及带千分位逗号或小数点的字符串。
    """
    # None 无法解析为整数，直接报错。
    if value is None:
        raise ValueError("Expected integer-like value, got None")
    # 已经是整数则原样返回。
    if isinstance(value, int):
        return value
    # 浮点数直接截断为整数。
    if isinstance(value, float):
        return int(value)
    # 其余按字符串处理：去空白、去千分位逗号。
    text = str(value).strip().replace(",", "")
    if not text:
        raise ValueError("Expected integer-like value, got empty string")
    # 先转 float 再转 int，可兼容 "12.0" 这类写法。
    return int(float(text))


def print_excel_preview(sheet_name: str, rows: list[dict[str, Any]]) -> None:
    """打印某个 Excel sheet 的预览信息(列名 + 前 3 行)，便于调试。"""
    print(f"[Excel预览] sheet={sheet_name}")
    # 空表时给出占位输出。
    if not rows:
        print("列名: []")
        print("前3行: []")
        return

    # 用第一行的键作为列名。
    columns = list(rows[0].keys())
    print(f"列名: {columns}")
    print("前3行:")
    # 仅打印前 3 行内容，ensure_ascii=False 以正常显示中文。
    for row in rows[:3]:
        print(json.dumps(row, ensure_ascii=False))


def load_segment_metadata(patient_name: str) -> list[SegmentMeta]:
    """从 Excel 中读取指定患者(sheet)的所有片段元数据。

    每个患者对应 Excel 中一个 sheet；逐行解析为 SegmentMeta 列表。
    """
    # 以只读、仅取值(不取公式)的方式打开工作簿，性能更好。
    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    try:
        # 患者名必须对应一个已存在的 sheet。
        if patient_name not in workbook.sheetnames:
            raise ValueError(f"Excel中不存在sheet: {patient_name}")

        sheet = workbook[patient_name]
        # 用迭代器逐行读取(values_only 表示只取单元格值)。
        rows_iter = sheet.iter_rows(values_only=True)
        # 第一行作为表头。
        header_row = next(rows_iter)
        if header_row is None:
            raise ValueError(f"sheet为空: {patient_name}")

        # 把表头规范化为字符串列表(None 视为空字符串)。
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        data_rows: list[dict[str, Any]] = []
        # 逐行把单元格值与表头组合成字典。
        for values in rows_iter:
            row = {headers[idx]: values[idx] for idx in range(len(headers))}
            # 跳过整行均为空的行。
            if all(value is None for value in row.values()):
                continue
            data_rows.append(row)

        # 打印预览，方便人工核对表结构是否正确。
        print_excel_preview(patient_name, data_rows)

        # 校验必须存在的列，缺一不可。
        required_columns = ["文件夹名称", "视频文件名", "开始帧号", "结束帧号"]
        missing_columns = [column for column in required_columns if column not in headers]
        if missing_columns:
            raise ValueError(f"Excel缺少必要列: {missing_columns}")

        metadata: list[SegmentMeta] = []
        for row in data_rows:
            session_name = str(row["文件夹名称"]).strip()
            video_name = str(row["视频文件名"]).strip()
            # 跳过会话名/视频名为空或显式为 "none" 的无效行。
            if not session_name or not video_name or video_name.lower() == "none":
                continue

            # 组装成不可变的 SegmentMeta 数据对象。
            metadata.append(
                SegmentMeta(
                    session_name=session_name,
                    video_name=video_name,
                    video_stem=Path(video_name).stem,
                    start_frame=parse_int(row["开始帧号"]),
                    end_frame=parse_int(row["结束帧号"]),
                )
            )

        return metadata
    finally:
        # 无论成功与否都关闭工作簿，释放文件句柄。
        workbook.close()


def extract_session_segment_info(segment_dir: Path) -> tuple[str, int, int]:
    """从片段目录名中解析出(归一化视频名, 起始帧, 结束帧)。

    目录名形如 ``<video>_<start>_<end>``，可能带 ``_skeleton_msk`` 后缀。
    """
    base_name = segment_dir.name
    # 去掉可能存在的 "_skeleton_msk" 后缀。
    if base_name.endswith("_skeleton_msk"):
        base_name = base_name[: -len("_skeleton_msk")]

    # 用正则提取视频名与首尾两个帧号(均为末尾的数字)。
    match = re.match(r"^(?P<video>.+)_(?P<start>\d+)_(?P<end>\d+)$", base_name)
    if not match:
        raise ValueError(f"无法从目录名解析视频与帧号: {segment_dir}")

    # 视频名归一化后返回，便于与 Excel 中的名称匹配。
    return (
        normalize_video_name(match.group("video")),
        int(match.group("start")),
        int(match.group("end")),
    )


def build_segment_json_index(patient_dir: Path) -> dict[tuple[str, str, int, int], SegmentSource]:
    """扫描患者目录下所有 all_frames_ankle_data.json，建立片段索引。

    索引键为 (会话名, 归一化视频名, 起始帧, 结束帧)，值为 SegmentSource。
    """
    index: dict[tuple[str, str, int, int], SegmentSource] = {}
    # 递归查找所有目标 JSON(排序保证结果稳定)。
    json_paths = sorted(patient_dir.glob("*/**/all_frames_ankle_data.json"))
    for json_path in json_paths:
        # JSON 的上上级目录名就是会话名。
        session_name = json_path.parent.parent.name
        # 从 JSON 所在目录名解析视频名与帧号。
        normalized_video, start_frame, end_frame = extract_session_segment_info(json_path.parent)
        index[(session_name, normalized_video, start_frame, end_frame)] = SegmentSource(
            json_path=json_path,
            segment_dir=json_path.parent,
        )
    return index


def print_json_preview(first_json_path: Path) -> None:
    """打印某个 JSON 文件的前两条记录预览(截断到 4000 字符)，便于调试。"""
    data = json.loads(first_json_path.read_text(encoding="utf-8"))
    # 若是列表则取前两条，否则原样预览。
    preview = data[:2] if isinstance(data, list) else data
    print(f"[JSON预览] {first_json_path}")
    print(json.dumps(preview, ensure_ascii=False, indent=2)[:4000])


def get_keypoint_xy(keypoints: Any, index: int) -> tuple[float, float] | None:
    """从关键点列表中安全地取出第 index 个点的 (x, y) 坐标。

    任何结构异常或坐标缺失都返回 None，避免后续计算崩溃。
    """
    # 必须是列表且索引在范围内。
    if not isinstance(keypoints, list) or index >= len(keypoints):
        return None

    point = keypoints[index]
    # 单个点必须是至少包含 2 个元素的列表。
    if not isinstance(point, list) or len(point) < 2:
        return None

    x = point[0]
    y = point[1]
    # 坐标值不能为 None。
    if x is None or y is None:
        return None

    # 尝试转为浮点数，失败则视为无效点。
    try:
        return float(x), float(y)
    except (TypeError, ValueError):
        return None


def midpoint(keypoints: Any, indices: tuple[int, int]) -> tuple[float, float] | None:
    """计算两个关键点的中点坐标；任一点缺失则返回 None。"""
    first = get_keypoint_xy(keypoints, indices[0])
    second = get_keypoint_xy(keypoints, indices[1])
    if first is None or second is None:
        return None
    # 分别对 x、y 取平均。
    return ((first[0] + second[0]) / 2.0, (first[1] + second[1]) / 2.0)


def segment_anchor(keypoints: Any, anchor: int | tuple[int, int]) -> tuple[float, float] | None:
    """获取某个段的锚点坐标。

    anchor 为元组时取中点，为单个索引时取该关键点本身。
    """
    if isinstance(anchor, tuple):
        return midpoint(keypoints, anchor)
    return get_keypoint_xy(keypoints, anchor)


def compute_frame_com(keypoints: Any) -> dict[str, float] | None:
    """用 7 段人体测量学模型计算单帧的质心(CoM)像素坐标。

    做法：对每个段，按近端->远端比例求该段质心位置，再以段质量为权重做加权平均。
    任一所需关键点缺失则整帧返回 None。
    """
    weighted_x = 0.0      # 累加的加权 x
    weighted_y = 0.0      # 累加的加权 y
    total_weight = 0.0    # 累加的总权重(质量比例之和)

    # 遍历每个身体段。
    for _, proximal_anchor, distal_anchor, weight, ratio in SEGMENTS:
        # 取该段的近端、远端锚点坐标。
        proximal = segment_anchor(keypoints, proximal_anchor)
        distal = segment_anchor(keypoints, distal_anchor)
        # 任一锚点缺失则无法计算该帧质心。
        if proximal is None or distal is None:
            return None

        # 段质心 = 近端 + ratio * (远端 - 近端)，即沿段方向按比例插值。
        segment_x = proximal[0] + ratio * (distal[0] - proximal[0])
        segment_y = proximal[1] + ratio * (distal[1] - proximal[1])

        # 以段质量(weight)为权重累加。
        weighted_x += weight * segment_x
        weighted_y += weight * segment_y
        total_weight += weight

    # 总权重为 0(理论上不会发生)时无法求平均。
    if total_weight == 0:
        return None

    # 加权平均得到全身质心，结果保留 4 位小数。
    return {
        "com_x": round(weighted_x / total_weight, 4),
        "com_y": round(weighted_y / total_weight, 4),
    }


def compute_segment_output(meta: SegmentMeta, json_path: Path) -> dict[str, Any]:
    """读取某片段的关键点 JSON，逐帧计算质心，组装为输出字典。"""
    # 读取该片段所有帧的关键点数据。
    frame_entries = json.loads(json_path.read_text(encoding="utf-8"))
    if not isinstance(frame_entries, list):
        raise ValueError(f"JSON根结构不是list: {json_path}")

    # frames: 帧号(字符串) -> 该帧质心(或 None)。
    frames: dict[str, dict[str, float] | None] = {}
    for entry in frame_entries:
        # 跳过非字典的异常项。
        if not isinstance(entry, dict):
            continue

        # 取帧号字段。
        frame_number = entry.get("frame")
        if frame_number is None:
            continue

        # 帧号解析失败的项跳过。
        try:
            frame_key = str(parse_int(frame_number))
        except ValueError:
            continue

        # 计算该帧质心(可能为 None)。
        frames[frame_key] = compute_frame_com(entry.get("keypoints"))

    # 返回包含视频名、起止帧与逐帧质心的结构。
    return {
        "video": meta.video_stem,
        "start_frame": meta.start_frame,
        "end_frame": meta.end_frame,
        "frames": frames,
    }


def draw_com_overlays(segment_dir: Path, frames: dict[str, dict[str, float] | None]) -> None:
    """在原始帧图上绘制质心圆点及坐标文字，输出到 com_overlay_frames/ 目录。"""
    # 绘图依赖 OpenCV，缺失时给出明确提示。
    if cv2 is None:
        raise RuntimeError("缺少cv2，请先安装opencv-python后再运行绘图功能")

    # 创建(或复用)叠加图输出目录。
    output_dir = segment_dir / "com_overlay_frames"
    output_dir.mkdir(exist_ok=True)

    # 逐帧处理。
    for frame_key, com in frames.items():
        # 原始帧图按 "frame_000123.jpg" 这种 6 位补零命名。
        source_image = segment_dir / f"frame_{int(frame_key):06d}.jpg"
        if not source_image.exists():
            continue

        # 读取图像，失败(返回 None)则跳过。
        image = cv2.imread(str(source_image))
        if image is None:
            continue

        # 仅在该帧成功算出质心时才绘制标记。
        if com is not None:
            center = (int(round(com["com_x"])), int(round(com["com_y"])))
            # 实心红点标记质心位置。
            cv2.circle(image, center, 6, (0, 0, 255), -1)
            # 外圈白色空心圆增强可见性。
            cv2.circle(image, center, 12, (255, 255, 255), 2)
            # 在质心旁标注精确坐标文字。
            cv2.putText(
                image,
                f"COM ({com['com_x']:.1f}, {com['com_y']:.1f})",
                (center[0] + 10, max(center[1] - 10, 25)),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.5,
                (0, 0, 255),
                1,
                cv2.LINE_AA,
            )

        # 以原文件名写出到叠加图目录。
        output_image = output_dir / source_image.name
        cv2.imwrite(str(output_image), image)


def main() -> int:
    """命令行入口：根据患者名处理其全部片段，输出 CoM JSON 与叠加图。"""
    # 必须且仅接受一个参数：患者名(即 Excel sheet 名)。
    if len(sys.argv) != 2:
        print("用法: python calculate_com.py 3_MsSu")
        return 1

    patient_name = sys.argv[1]
    patient_dir = CAMERA_DATA_ROOT / patient_name
    # 患者目录不存在则直接退出。
    if not patient_dir.exists():
        print(f"患者目录不存在: {patient_dir}")
        return 1

    # 读取 Excel 元数据，并扫描磁盘建立片段索引。
    metadata = load_segment_metadata(patient_name)
    segment_index = build_segment_json_index(patient_dir)
    # 若有片段，打印第一个 JSON 的预览供核对。
    if segment_index:
        first_segment_source = next(iter(segment_index.values()))
        print_json_preview(first_segment_source.json_path)

    missing_segments: list[str] = []
    # 逐条元数据去索引里找对应的磁盘片段。
    for meta in metadata:
        normalized_video = normalize_video_name(meta.video_stem)
        segment_source = segment_index.get((meta.session_name, normalized_video, meta.start_frame, meta.end_frame))
        # 找不到则记录为缺失片段，稍后统一报告。
        if segment_source is None:
            missing_segments.append(
                f"{patient_name}/{meta.session_name}/{meta.video_stem}_{meta.start_frame}_{meta.end_frame}"
            )
            continue

        # 计算该片段逐帧质心。
        output = compute_segment_output(meta, segment_source.json_path)
        # 输出 JSON 文件名与路径(写在会话目录下)。
        output_name = f"{meta.video_stem}_{meta.start_frame}_{meta.end_frame}_COM.json"
        output_path = patient_dir / meta.session_name / output_name
        output_path.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
        # 同时绘制质心叠加图。
        draw_com_overlays(segment_source.segment_dir, output["frames"])
        print(f"[完成] {patient_name} / {meta.session_name} / {meta.video_stem}_{meta.start_frame}_{meta.end_frame} → {output_name}")

    # 统一报告所有未匹配到磁盘数据的片段。
    if missing_segments:
        print("[警告] 以下segment未找到对应的all_frames_ankle_data.json:")
        for item in missing_segments:
            print(f"  - {item}")

    return 0


if __name__ == "__main__":
    # 以 main() 的返回值作为进程退出码。
    raise SystemExit(main())
